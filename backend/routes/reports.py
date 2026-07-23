import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ReportLab Imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from database import get_db
from models import Product, Inventory, Sale, User
from routes.utils import get_current_user

reports_router = APIRouter()

# --- SALES CSV ---
@reports_router.get('/sales/csv')
def export_sales_csv(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    sales = db.query(Sale).order_by(Sale.sale_date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['Sale ID', 'Sale Date', 'Product ID', 'Product Name', 'Quantity', 'Sale Price', 'Total Amount'])
    
    for s in sales:
        writer.writerow([
            s.id,
            s.sale_date.strftime('%Y-%m-%d %H:%M:%S') if s.sale_date else '',
            s.product_id,
            s.product.name if s.product else 'Unknown',
            s.quantity,
            s.sale_price,
            s.total_amount
        ])
        
    output.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="sales_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv"'
    }
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/csv',
        headers=headers
    )

# --- SALES EXCEL ---
@reports_router.get('/sales/excel')
def export_sales_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    sales = db.query(Sale).order_by(Sale.sale_date.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "SmartKirana Sales Report"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Subheader row
    ws.append([]) # Blank row
    
    # Header fields
    headers = ['Sale ID', 'Sale Date', 'Product ID', 'Product Name', 'Quantity', 'Sale Price', 'Total Amount']
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True)
    for col_idx in range(1, 8):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Append data
    total_rev = 0
    total_qty = 0
    for s in sales:
        total_rev += s.total_amount
        total_qty += s.quantity
        ws.append([
            s.id,
            s.sale_date.strftime('%Y-%m-%d %H:%M:%S') if s.sale_date else '',
            s.product_id,
            s.product.name if s.product else 'Unknown',
            s.quantity,
            s.sale_price,
            s.total_amount
        ])
        
    # Total row
    ws.append([])
    tot_row = len(sales) + 5
    ws.cell(row=tot_row, column=4, value="TOTALS").font = Font(bold=True)
    ws.cell(row=tot_row, column=5, value=total_qty).font = Font(bold=True)
    ws.cell(row=tot_row, column=7, value=total_rev).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in ws.columns:
        # Check if values exist in column
        vals = [cell.value for cell in col if cell.value is not None]
        if vals:
            max_len = max(len(str(v)) for v in vals)
        else:
            max_len = 0
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="sales_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    }
    return StreamingResponse(
        out,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

# --- SALES PDF ---
@reports_router.get('/sales/pdf')
def export_sales_pdf(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    sales = db.query(Sale).order_by(Sale.sale_date.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter, 
        rightMargin=36, 
        leftMargin=36, 
        topMargin=36, 
        bottomMargin=36
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=8,
        spaceBefore=10
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=15
    )
    
    # Document Header
    story.append(Paragraph("SmartKirana - Sales Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Summary Dashboard Table
    total_rev = sum(s.total_amount for s in sales)
    total_qty = sum(s.quantity for s in sales)
    
    summary_data = [
        ["Total Revenue", "Total Quantity Sold", "Total Transactions"],
        [f"INR {total_rev:,.2f}", f"{total_qty:,}", f"{len(sales):,}"]
    ]
    summary_table = Table(summary_data, colWidths=[180, 180, 180])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ECEFF1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#37474F')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CFD8DC')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1F4E79')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Main Sales Table
    table_data = [["ID", "Date/Time", "Product Name", "Qty", "Price (INR)", "Total (INR)"]]
    for s in sales[:100]: # Limit to 100 entries for PDF size safety
        table_data.append([
            str(s.id),
            s.sale_date.strftime('%Y-%m-%d %H:%M') if s.sale_date else '',
            s.product.name if s.product else 'Unknown',
            str(s.quantity),
            f"{s.sale_price:.2f}",
            f"{s.total_amount:.2f}"
        ])
        
    # Col widths (Total 540 for printable region)
    col_widths = [40, 110, 200, 50, 70, 70]
    sales_table = Table(table_data, colWidths=col_widths)
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),  # Left align product name
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B0BEC5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(sales_table)
    
    # If truncated message
    if len(sales) > 100:
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<i>* Showing first 100 transactions of {len(sales)}. For full dataset export to Excel/CSV.</i>", subtitle_style))
        
    doc.build(story)
    buffer.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="sales_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.pdf"'
    }
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers=headers
    )

# --- INVENTORY STATUS EXCEL ---
@reports_router.get('/inventory/excel')
def export_inventory_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    inventory_items = db.query(Inventory).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "SmartKirana Inventory Report"
    title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid') # Green theme
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.append([]) # blank
    
    headers = ['Product ID', 'Product Name', 'Category', 'Current Stock', 'Min Stock Level', 'Status']
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    header_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True)
    for col_idx in range(1, 7):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    for item in inventory_items:
        p = item.product
        status_str = "Normal"
        if item.current_stock <= 0:
            status_str = "OUT OF STOCK"
        elif item.current_stock <= p.min_stock_level:
            status_str = "LOW STOCK"
            
        ws.append([
            p.id,
            p.name,
            p.category,
            item.current_stock,
            p.min_stock_level,
            status_str
        ])
        
        # Color specific status
        current_row = ws.max_row
        status_cell = ws.cell(row=current_row, column=6)
        if status_str == "OUT OF STOCK":
            status_cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            status_cell.font = Font(color='B71C1C', bold=True)
        elif status_str == "LOW STOCK":
            status_cell.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
            status_cell.font = Font(color='E65100', bold=True)
        else:
            status_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            status_cell.font = Font(color='1B5E20')
            
    # Auto-adjust column widths
    for col in ws.columns:
        vals = [cell.value for cell in col if cell.value is not None]
        if vals:
            max_len = max(len(str(v)) for v in vals)
        else:
            max_len = 0
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="inventory_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    }
    return StreamingResponse(
        out,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )
